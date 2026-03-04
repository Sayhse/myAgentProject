# @Time    : 2026/3/2 15:47
# @Author  : Yun
# @FileName: ui_use_case_code_generator_graph
# @Software: PyCharm
# @Desc    :
from langchain.chat_models import init_chat_model
from langchain_core.language_models import BaseChatModel
from langchain_core.messages import SystemMessage
from langgraph.checkpoint.memory import InMemorySaver
from langgraph.config import get_stream_writer
from langgraph.constants import START, END
from langgraph.graph import StateGraph

from auto_test_assistant.agents.code_generator_agent import CodeGeneratorAgentSystem
from auto_test_assistant.state.ui_usecase_code_generator_state import UiUseCaseCodeGeneratorState
import openpyxl
from pathlib import Path


def build_llm() -> BaseChatModel:
    """
    构建底层 LLM。
    默认使用 OpenAI 风格接口，你可以根据自己环境替换为其他实现（如自建 LLM 网关）。
    需要环境变量：
    - OPENAI_API_KEY
    - OPENAI_MODEL
    """
    llm = init_chat_model(model="deepseek-chat", model_provider="deepseek", max_tokens=8192)
    return llm


def use_case_splitting(state: UiUseCaseCodeGeneratorState):
    print(">>> use_case_splitting_node")
    writer = get_stream_writer()
    writer({"node": ">>> use_case_splitting_node"})

    uploaded_files_metadata = state.get("uploaded_files_metadata", [])
    ui_use_cases = state.get("ui_use_cases", [])

    required_columns = ["用例编号", "功能描述", "前置条件", "测试步骤", "预期结果", "后置条件"]

    def split_test_steps(steps_str):
        if not steps_str:
            return []

        steps_list = []
        lines = str(steps_str).strip().split('\n')

        for line in lines:
            line = line.strip()
            if not line:
                continue

            n = None
            value = line

            import re
            match = re.match(r'^(\d+)\.\s*(.*)', line)
            if match:
                n = int(match.group(1))
                value = match.group(2).strip()
            else:
                if not steps_list:
                    n = 1
                else:
                    n = steps_list[-1]["id"] + 1

            if value.endswith(';'):
                value = value[:-1]

            steps_list.append({"id": n, "value": value})

        return steps_list

    for file_meta in uploaded_files_metadata:
        file_path = file_meta.get("path")
        if not file_path:
            continue

        path = Path(file_path)
        if path.suffix.lower() not in [".xlsx", ".xls"]:
            continue

        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            writer({"error": f"无法加载Excel文件 {file_path}: {e}"})
            continue

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]

            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header_row = row
                break

            if not header_row:
                continue

            column_indices = {}
            for idx, cell_value in enumerate(header_row, start=1):
                if cell_value is None:
                    continue
                column_indices[str(cell_value).strip()] = idx

            missing_columns = [col for col in required_columns if col not in column_indices]
            if missing_columns:
                writer({"warning": f"Sheet {sheet_name} 缺少必需列: {missing_columns}"})
                continue

            for row in ws.iter_rows(min_row=2, values_only=True):
                case_id = row[column_indices["用例编号"] - 1] if column_indices["用例编号"] <= len(row) else None
                if not case_id:
                    continue

                use_case = {"module": sheet_name}
                for col in required_columns:
                    col_idx = column_indices[col]
                    value = row[col_idx - 1] if col_idx <= len(row) else None
                    if col == "测试步骤":
                        use_case[col] = split_test_steps(value)
                    else:
                        use_case[col] = value

                ui_use_cases.append(use_case)

    return {"messages": [SystemMessage(content="用例拆分完成")], "type": "use_case_splitting",
            "ui_use_cases": ui_use_cases}


def code_generator(state: UiUseCaseCodeGeneratorState):
    print(">>> code_generator_node")
    writer = get_stream_writer()
    writer({"node": ">>> code_generator_node"})
    use_cases = state.get("ui_use_cases", [])
    for use_case in use_cases:
        llm = build_llm()
        system = CodeGeneratorAgentSystem(llm=llm)
        result = system.code_generation(use_case)
    return {"messages": [SystemMessage(content="code_generator")], "type": "code_generator",
            "reason": "该问题与软件测试无关"}


def code_review(state: UiUseCaseCodeGeneratorState):
    print(">>> code_review_node")
    writer = get_stream_writer()
    writer({"node": ">>> code_review_node"})
    return {"messages": [SystemMessage(content="code_review")], "type": "code_review",
            "reason": "该问题与软件测试无关"}


def build_code_generator_agent():
    """
    根据UI测试用例生成代码的graph
    """
    builder = StateGraph(UiUseCaseCodeGeneratorState)
    builder.add_node("use_case_splitting", use_case_splitting)
    builder.add_node("code_generator", code_generator)
    builder.add_node("code_review", code_review)

    builder.add_edge(START, "use_case_splitting")
    builder.add_edge("use_case_splitting", "code_generator")
    builder.add_edge("code_generator", "code_review")
    builder.add_edge("code_review", END)

    checkpoint = InMemorySaver()
    graph = builder.compile(checkpointer=checkpoint)
    return graph
